import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from collections import defaultdict


def split_by_duplicate(
        input_file: str = "output/merged_clean.xlsx",
        unique_output: str = "output/unique.xlsx",
        duplicate_first_output: str = "output/duplicates_first.xlsx",
        duplicate_second_output: str = "output/duplicates_second.xlsx"):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    headers = None
    data = []

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = row
            continue
        data.append(row)

    wb.close()

    if not data:
        print("警告: 文件中没有数据")
        return

    training_name_col_idx = None
    for i, header in enumerate(headers):
        if header and "培训/会议名称" in header:
            training_name_col_idx = i
            break

    if training_name_col_idx is None:
        print("警告: 未找到'培训/会议名称'列")
        return

    name_occurrences = defaultdict(list)
    for row_idx, row in enumerate(data):
        name = row[training_name_col_idx]
        if name:
            name_occurrences[name].append((row_idx, row))

    unique_data = []
    duplicate_first_data = []
    duplicate_second_data = []

    for name, occurrences in name_occurrences.items():
        if len(occurrences) == 1:
            unique_data.append(occurrences[0][1])
        else:
            for idx, (row_idx, row) in enumerate(occurrences):
                if idx == 0:
                    duplicate_first_data.append(row)
                elif idx == 1:
                    duplicate_second_data.append(row)
                else:
                    duplicate_second_data.append(row)

    def save_to_workbook(data_rows, output_path, sheet_title):
        if not data_rows:
            print(f"警告: 没有数据写入 {output_path}")
            return

        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        output_ws.title = sheet_title

        header_fill = PatternFill(start_color="4472C4",
                                  end_color="4472C4",
                                  fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, start=1):
            cell = output_ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                output_ws.cell(row=row_idx, column=col_idx, value=value)

        for column in output_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            output_ws.column_dimensions[column_letter].width = adjusted_width

        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        output_wb.save(output_path)
        print(f"已保存: {output_path}")

    save_to_workbook(unique_data, unique_output, "不重复")
    save_to_workbook(duplicate_first_data, duplicate_first_output, "重复-第一行")
    save_to_workbook(duplicate_second_data, duplicate_second_output, "重复-第二行+")

    print()
    print("=" * 50)
    print("拆分结果汇总")
    print("=" * 50)
    print(f"总记录数: {len(data)}")
    print(f"不重复记录: {len(unique_data)}")
    print(f"重复记录-第一行: {len(duplicate_first_data)}")
    print(f"重复记录-第二行及更多: {len(duplicate_second_data)}")
    print(
        f"重复的培训/会议名称数: {len([name for name, occ in name_occurrences.items() if len(occ) > 1])}"
    )
    print("=" * 50)


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='按培训/会议名称重复情况拆分Excel文件')
    parser.add_argument('-i',
                        '--input',
                        default='output/merged_clean.xlsx',
                        help='输入文件路径（默认: output/merged_clean.xlsx）')
    parser.add_argument('-u',
                        '--unique',
                        default='output/unique.xlsx',
                        help='不重复记录输出文件（默认: output/unique.xlsx）')
    parser.add_argument('-f',
                        '--first',
                        default='output/duplicates_first.xlsx',
                        help='重复记录第一行输出文件（默认: output/duplicates_first.xlsx）')
    parser.add_argument('-s',
                        '--second',
                        default='output/duplicates_second.xlsx',
                        help='重复记录第二行输出文件（默认: output/duplicates_second.xlsx）')

    args = parser.parse_args()

    print("=" * 50)
    print("Excel按重复拆分工具")
    print("=" * 50)
    print(f"输入文件: {args.input}")
    print(f"不重复输出: {args.unique}")
    print(f"重复-第一行输出: {args.first}")
    print(f"重复-第二行+输出: {args.second}")
    print("=" * 50)
    print()

    split_by_duplicate(args.input, args.unique, args.first, args.second)
