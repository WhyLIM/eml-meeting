import os
import sys
import argparse
import csv
from datetime import datetime
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import glob

from eml_parser import parse_eml_file
from extractor import extract_training_info_batch
from config import DEFAULT_API, get_available_apis
from config_loader import ConfigLoader


def find_eml_files(input_dir: str) -> List[str]:
    eml_files = []

    if os.path.isfile(input_dir):
        if input_dir.lower().endswith('.eml'):
            eml_files.append(input_dir)
    elif os.path.isdir(input_dir):
        eml_files.extend(glob.glob(os.path.join(input_dir, '*.eml')))
        eml_files.extend(
            glob.glob(os.path.join(input_dir, '**/*.eml'), recursive=True))

    return sorted(eml_files)


def parse_all_emls(eml_files: List[str]) -> List[dict]:
    parsed_data = []
    total = len(eml_files)

    print(f"\n开始解析 {total} 个EML文件...")

    for i, file_path in enumerate(eml_files):
        try:
            print(f"[{i+1}/{total}] 解析: {os.path.basename(file_path)}")
            data = parse_eml_file(file_path)
            parsed_data.append(data)
        except Exception as e:
            print(f"  警告: 解析失败 - {e}")
            parsed_data.append({
                'file_path': file_path,
                'file_name': os.path.basename(file_path),
                'subject': '',
                'from': '',
                'date': '',
                'body': '',
                'error': str(e)
            })

    print(f"完成！成功解析 {len(parsed_data)} 个文件")
    return parsed_data


def print_progress(current: int, total: int, filename: str):
    percentage = (current / total) * 100
    print(
        f"\r[{'='*int(percentage/5):<20}] {current}/{total} ({percentage:.1f}%) - {filename[:30]}",
        end='',
        flush=True)
    if current == total:
        print()


def save_to_csv(results: List[dict], output_path: str):
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    fieldnames = [
        '文件名', '培训/会议名称', '开始时间', '结束时间', '学时(小时)', '地点', '讲座目的', '讲座内容',
        '提取状态'
    ]

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        for result in results:
            status = '成功' if result.get('training_name') else (
                '失败: ' + result.get('error', '未知错误'))

            writer.writerow({
                '文件名': result.get('file_name', ''),
                '培训/会议名称': result.get('training_name', ''),
                '开始时间': result.get('start_time', ''),
                '结束时间': result.get('end_time', ''),
                '学时(小时)': result.get('duration_hours', ''),
                '地点': result.get('location', ''),
                '讲座目的': result.get('purpose', ''),
                '讲座内容': result.get('content', ''),
                '提取状态': status
            })

    print(f"\nCSV文件已保存: {output_path}")


def save_to_excel(results: List[dict], output_path: str):
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学术报告信息"

    headers = [
        '文件名', '培训/会议名称', '开始时间', '结束时间', '学时(小时)', '地点', '讲座目的', '讲座内容',
        '提取状态'
    ]

    header_fill = PatternFill(start_color="4472C4",
                              end_color="4472C4",
                              fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row, result in enumerate(results, start=2):
        status = '成功' if result.get('training_name') else (
            '失败: ' + result.get('error', '未知错误'))

        ws.cell(row=row, column=1, value=result.get('file_name', ''))
        ws.cell(row=row, column=2, value=result.get('training_name', ''))
        ws.cell(row=row, column=3, value=result.get('start_time', ''))
        ws.cell(row=row, column=4, value=result.get('end_time', ''))
        ws.cell(row=row, column=5, value=result.get('duration_hours', ''))
        ws.cell(row=row, column=6, value=result.get('location', ''))
        ws.cell(row=row, column=7, value=result.get('purpose', ''))
        ws.cell(row=row, column=8, value=result.get('content', ''))
        ws.cell(row=row, column=9, value=status)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"\nExcel文件已保存: {output_path}")


def print_summary(results: List[dict]):
    total_records = len(results)
    success = sum(1 for r in results if r.get('training_name'))
    failed = total_records - success

    files_with_errors = {}
    for r in results:
        if not r.get('training_name'):
            fname = r.get('file_name', 'unknown')
            files_with_errors[fname] = r.get('error', '未知错误')

    print("\n" + "=" * 50)
    print("处理结果汇总")
    print("=" * 50)
    print(f"总记录数: {total_records}")
    print(f"成功提取: {success}")
    print(f"提取失败: {failed}")
    print("=" * 50)

    if failed > 0:
        print("\n失败记录:")
        for fname, error in files_with_errors.items():
            print(f"  - {fname}: {error}")


def main():
    parser = argparse.ArgumentParser(
        description='EML邮件学术报告信息提取工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python main.py --config config.yaml
  python main.py -i messages_package -o output/result.xlsx
  python main.py -i messages_package -o output/result.csv --api deepseek
  python main.py -i messages_package -o output/result.xlsx --api openai --model gpt-4-turbo
        """)

    parser.add_argument('-c',
                        '--config',
                        default='config.yaml',
                        help='配置文件路径（默认: config.yaml）')
    parser.add_argument('-i', '--input', help='输入EML文件或目录路径（覆盖配置文件）')
    parser.add_argument('-o', '--output', help='输出文件路径（覆盖配置文件）')
    parser.add_argument('--api',
                        choices=get_available_apis(),
                        help='选择LLM API提供商（覆盖配置文件）')
    parser.add_argument('--model', help='指定模型名称（覆盖配置文件）')

    args = parser.parse_args()

    config_loader = None
    if os.path.exists(args.config):
        try:
            config_loader = ConfigLoader(args.config)
            if not config_loader.validate():
                print(f"警告: 配置文件 {args.config} 缺少必要参数")
        except Exception as e:
            print(f"警告: 无法加载配置文件 {args.config}: {e}")
    else:
        print(f"警告: 配置文件 {args.config} 不存在")

    input_dir = args.input
    output_file = args.output
    api_provider = args.api
    model = args.model

    if config_loader:
        if not input_dir:
            input_dir = config_loader.get_input_dir()
        if not output_file:
            output_file = config_loader.get_output_file()
        if not api_provider:
            api_provider = config_loader.get_api_provider()
        if not model:
            model = config_loader.get_model()

    if not input_dir:
        print("错误: 未指定输入目录，请通过 -i 参数或配置文件指定")
        sys.exit(1)

    if not output_file:
        print("错误: 未指定输出文件，请通过 -o 参数或配置文件指定")
        sys.exit(1)

    if not api_provider:
        api_provider = DEFAULT_API

    if not os.path.exists(input_dir):
        print(f"错误: 输入路径不存在: {input_dir}")
        sys.exit(1)

    output_ext = os.path.splitext(output_file)[1].lower()
    if output_ext not in ['.xlsx', '.csv']:
        print("错误: 输出文件必须为 .xlsx 或 .csv 格式")
        sys.exit(1)

    print("=" * 50)
    print("EML邮件学术报告信息提取工具")
    print("=" * 50)
    print(f"配置文件: {args.config}")
    print(f"输入路径: {input_dir}")
    print(f"输出路径: {output_file}")
    print(f"API提供商: {api_provider}")
    if model:
        print(f"模型: {model}")
    print("=" * 50)

    eml_files = find_eml_files(input_dir)

    if not eml_files:
        print(f"警告: 未找到EML文件: {input_dir}")
        sys.exit(0)

    print(f"\n找到 {len(eml_files)} 个EML文件")

    parsed_data = parse_all_emls(eml_files)

    print("\n开始提取学术报告信息...")
    results = extract_training_info_batch(parsed_data, api_provider,
                                          print_progress)

    if output_ext == '.xlsx':
        save_to_excel(results, output_file)
    else:
        save_to_csv(results, output_file)

    print_summary(results)

    success_count = sum(1 for r in results if r.get('training_name'))
    if success_count == 0:
        print("\n警告: 所有文件提取失败，请检查API密钥配置和网络连接")


if __name__ == '__main__':
    main()
