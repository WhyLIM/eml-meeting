import os
import glob
import openpyxl
from datetime import datetime


def merge_excel_files(input_dir: str = "output",
                      output_file: str = "output/merged.xlsx"):
    xlsx_files = sorted(glob.glob(os.path.join(input_dir, "result*.xlsx")))

    if not xlsx_files:
        print(f"警告: 未找到Excel文件在 {input_dir}")
        return

    print(f"找到 {len(xlsx_files)} 个Excel文件")

    all_data = []

    for file_path in xlsx_files:
        file_name = os.path.basename(file_path)
        print(f"读取: {file_name}")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = row
                continue
            all_data.append(row)

        wb.close()

    if not all_data:
        print("警告: 没有数据可以合并")
        return

    merged_wb = openpyxl.Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "合并结果"

    header_fill = openpyxl.styles.PatternFill(start_color="4472C4",
                                              end_color="4472C4",
                                              fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = merged_ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal='center',
                                                   vertical='center')

    for row_idx, row_data in enumerate(all_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            merged_ws.cell(row=row_idx, column=col_idx, value=value)

    for column in merged_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        merged_ws.column_dimensions[column_letter].width = adjusted_width

    os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
    merged_wb.save(output_file)

    print(f"\n合并完成！")
    print(f"总记录数: {len(all_data)}")
    print(f"输出文件: {output_file}")


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='合并多个Excel文件')
    parser.add_argument('-i',
                        '--input',
                        default='output',
                        help='输入目录（默认: output）')
    parser.add_argument('-o',
                        '--output',
                        default='output/merged.xlsx',
                        help='输出文件路径（默认: output/merged.xlsx）')

    args = parser.parse_args()

    print("=" * 50)
    print("Excel文件合并工具")
    print("=" * 50)
    print(f"输入目录: {args.input}")
    print(f"输出文件: {args.output}")
    print("=" * 50)
    print()

    merge_excel_files(args.input, args.output)
